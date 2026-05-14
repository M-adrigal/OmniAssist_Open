"""
Excel 电子表格格式化引擎

将结构化的表格描述（JSON）转换为 openpyxl 工作簿。
支持多 Sheet、单元格样式、数字格式、行列操作、合并、条件格式、图表。

设计原则：
- 数据与样式分离：data 存原始值，styles 独立描述格式
- 范围引用支持 A1 表示法和列名引用（如"销售额列"）
- 未指定的属性使用合理默认值
- 无法精确实现的样式自动降级并在结果中提示
"""

import os
import re

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.formatting.rule import (
    CellIsRule, ColorScaleRule, DataBarRule, IconSetRule, FormulaRule
)
from openpyxl.chart import (
    BarChart, LineChart, PieChart, Reference, BarChart3D, PieChart3D
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension

DEFAULT_FONT = {
    "name": "宋体",
    "size": 11,
    "color": "#000000",
    "bold": False,
    "italic": False,
    "underline": "none",
}

DEFAULT_FILL = {
    "color": None,
}

DEFAULT_ALIGNMENT = {
    "horizontal": "left",
    "vertical": "center",
    "wrap_text": False,
    "indent": 0,
}

DEFAULT_BORDER = {
    "style": "thin",
    "color": "#000000",
}

DEFAULT_NUMBER_FORMAT = {
    "general": "General",
    "integer": "0",
    "decimal_1": "0.0",
    "decimal_2": "0.00",
    "thousands": "#,##0",
    "thousands_2": "#,##0.00",
    "currency": "¥#,##0.00",
    "currency_usd": "$#,##0.00",
    "percent": "0.00%",
    "percent_int": "0%",
    "date": "yyyy-mm-dd",
    "date_cn": "yyyy\"年\"m\"月\"d\"日\"",
    "datetime": "yyyy-mm-dd h:mm:ss",
    "time": "h:mm:ss",
    "text": "@",
    "scientific": "0.00E+00",
}

BORDER_STYLE_MAP = {
    "thin": "thin",
    "medium": "medium",
    "thick": "thick",
    "double": "double",
    "dashed": "dashed",
    "dotted": "dotted",
    "hair": "hair",
    "none": None,
}

HORIZONTAL_ALIGN_MAP = {
    "left": "left",
    "center": "center",
    "right": "right",
    "justify": "justify",
    "fill": "fill",
    "distributed": "distributed",
}

VERTICAL_ALIGN_MAP = {
    "top": "top",
    "center": "center",
    "bottom": "bottom",
    "justify": "justify",
    "distributed": "distributed",
}

UNDERLINE_MAP = {
    "none": "none",
    "single": "single",
    "double": "double",
    "singleAccounting": "singleAccounting",
    "doubleAccounting": "doubleAccounting",
}

CHART_TYPE_MAP = {
    "bar": ("bar", BarChart),
    "bar3d": ("bar3d", BarChart3D),
    "line": ("line", LineChart),
    "pie": ("pie", PieChart),
    "pie3d": ("pie3d", PieChart3D),
    "column": ("col", BarChart),
}

CHART_ALIAS_MAP = {
    "柱状图": "bar",
    "条形图": "bar",
    "折线图": "line",
    "饼图": "pie",
    "饼状图": "pie",
    "三维柱状图": "bar3d",
    "三维饼图": "pie3d",
    "柱形图": "bar",
    "直方图": "bar",
}

CONDITIONAL_OPERATOR_MAP = {
    "greaterThan": "greaterThan",
    "greater_than": "greaterThan",
    "大于": "greaterThan",
    "lessThan": "lessThan",
    "less_than": "lessThan",
    "小于": "lessThan",
    "equal": "equal",
    "等于": "equal",
    "notEqual": "notEqual",
    "not_equal": "notEqual",
    "不等于": "notEqual",
    "between": "between",
    "介于": "between",
    "notBetween": "notBetween",
    "不介于": "notBetween",
    "greaterThanOrEqual": "greaterThanOrEqual",
    "大于等于": "greaterThanOrEqual",
    "lessThanOrEqual": "lessThanOrEqual",
    "小于等于": "lessThanOrEqual",
}

ICON_SET_MAP = {
    "3Arrows": "3Arrows",
    "3TrafficLights1": "3TrafficLights1",
    "3TrafficLights2": "3TrafficLights2",
    "3Signs": "3Signs",
    "3Symbols": "3Symbols",
    "3Flags": "3Flags",
    "3Stars": "3Stars",
    "4Arrows": "4Arrows",
    "4Rating": "4Rating",
    "5Arrows": "5Arrows",
    "5Rating": "5Rating",
    "5Quarters": "5Quarters",
    "箭头": "3Arrows",
    "红绿灯": "3TrafficLights1",
    "信号灯": "3TrafficLights1",
    "星标": "3Stars",
    "星级": "3Stars",
    "旗帜": "3Flags",
}


def _parse_color(hex_color):
    """将 #RRGGBB 颜色转为 openpyxl 可用的字符串"""
    if not hex_color or not isinstance(hex_color, str):
        return None
    hex_color = hex_color.lstrip("#")
    if len(hex_color) == 6:
        return f"FF{hex_color}"
    if len(hex_color) == 8:
        return hex_color
    return None


def _parse_range(range_str):
    """解析范围字符串如 'A1:C10' 或 'A:A' 或 '1:10'

    Returns:
        (min_col, min_row, max_col, max_row) 或 None
    """
    if not range_str or not isinstance(range_str, str):
        return None

    range_str = range_str.strip().upper()

    # 整列: A:A 或 A:C
    col_match = re.match(r'^([A-Z]+):([A-Z]+)$', range_str)
    if col_match:
        c1 = column_index_from_string(col_match.group(1))
        c2 = column_index_from_string(col_match.group(2))
        return (c1, 1, c2, 1048576)

    # 整行: 1:10
    row_match = re.match(r'^(\d+):(\d+)$', range_str)
    if row_match:
        return (1, int(row_match.group(1)), 16384, int(row_match.group(2)))

    # 标准范围: A1:C10
    std_match = re.match(r'^([A-Z]+)(\d+):([A-Z]+)(\d+)$', range_str)
    if std_match:
        c1 = column_index_from_string(std_match.group(1))
        r1 = int(std_match.group(2))
        c2 = column_index_from_string(std_match.group(3))
        r2 = int(std_match.group(4))
        return (c1, r1, c2, r2)

    # 单格: A1
    single_match = re.match(r'^([A-Z]+)(\d+)$', range_str)
    if single_match:
        c = column_index_from_string(single_match.group(1))
        r = int(single_match.group(2))
        return (c, r, c, r)

    return None


def _resolve_range(range_spec, ws, warnings):
    """解析范围描述，支持 A1 表示法和列名引用

    Args:
        range_spec: 范围描述，可以是 A1 字符串或 {"column": "销售额", "start_row": 2} 等
        ws: 工作表对象
        warnings: 警告列表

    Returns:
        (min_col, min_row, max_col, max_row) 或 None
    """
    if isinstance(range_spec, str):
        return _parse_range(range_spec)

    if isinstance(range_spec, dict):
        col_name = range_spec.get("column")
        if col_name:
            col_idx = _find_column_by_header(ws, col_name)
            if col_idx is None:
                warnings.append(f"未找到列 '{col_name}'，跳过样式应用")
                return None
            start_row = range_spec.get("start_row", 2)
            end_row = range_spec.get("end_row", ws.max_row)
            return (col_idx, start_row, col_idx, end_row)

        col_start = range_spec.get("col_start")
        col_end = range_spec.get("col_end")
        row_start = range_spec.get("row_start", 1)
        row_end = range_spec.get("row_end")

        if col_start:
            if isinstance(col_start, str) and not col_start[0].isdigit():
                col_start = _find_column_by_header(ws, col_start)
            if col_start is None:
                return None
        if col_end:
            if isinstance(col_end, str) and not col_end[0].isdigit():
                col_end = _find_column_by_header(ws, col_end)
            if col_end is None:
                return None

        if col_start is None:
            return None

        return (
            col_start,
            row_start,
            col_end or col_start,
            row_end or ws.max_row,
        )

    return None


def _find_column_by_header(ws, header_name):
    """在表头行中查找列名对应的列索引"""
    if ws.max_row < 1:
        return None
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value and str(cell_value).strip() == header_name.strip():
            return col
    return None


def _build_font(font_spec, base_font=None):
    """构建 openpyxl Font 对象"""
    merged = {}
    if base_font:
        merged.update(base_font)
    if font_spec:
        merged.update(font_spec)

    kwargs = {}
    if "name" in merged:
        kwargs["name"] = merged["name"]
    if "size" in merged:
        kwargs["size"] = merged["size"]
    color = _parse_color(merged.get("color"))
    if color:
        kwargs["color"] = color
    if merged.get("bold"):
        kwargs["bold"] = True
    if merged.get("italic"):
        kwargs["italic"] = True
    underline = merged.get("underline", "none")
    if underline in UNDERLINE_MAP and underline != "none":
        kwargs["underline"] = UNDERLINE_MAP[underline]

    return Font(**kwargs) if kwargs else None


def _build_fill(fill_spec):
    """构建 openpyxl PatternFill 对象"""
    if not fill_spec:
        return None

    bg_color = fill_spec.get("color") or fill_spec.get("bg_color")
    if not bg_color:
        return None

    color = _parse_color(bg_color)
    if not color:
        return None

    pattern = fill_spec.get("pattern", "solid")
    return PatternFill(start_color=color, end_color=color, fill_type=pattern)


def _build_alignment(align_spec, base_align=None):
    """构建 openpyxl Alignment 对象"""
    merged = {}
    if base_align:
        merged.update(base_align)
    if align_spec:
        merged.update(align_spec)

    kwargs = {}
    h = merged.get("horizontal")
    if h and h in HORIZONTAL_ALIGN_MAP:
        kwargs["horizontal"] = HORIZONTAL_ALIGN_MAP[h]
    v = merged.get("vertical")
    if v and v in VERTICAL_ALIGN_MAP:
        kwargs["vertical"] = VERTICAL_ALIGN_MAP[v]
    if "wrap_text" in merged:
        kwargs["wrap_text"] = merged["wrap_text"]
    if "indent" in merged:
        kwargs["indent"] = merged["indent"]

    return Alignment(**kwargs) if kwargs else None


def _build_border(border_spec):
    """构建 openpyxl Border 对象"""
    if not border_spec:
        return None

    style = border_spec.get("style", "thin")
    color_str = border_spec.get("color", "#000000")
    color = _parse_color(color_str) or "FF000000"

    side_style = BORDER_STYLE_MAP.get(style, "thin")
    if side_style is None:
        return None

    side = Side(style=side_style, color=color)

    # 支持分别设置各边
    top = _build_side(border_spec.get("top"), side)
    bottom = _build_side(border_spec.get("bottom"), side)
    left = _build_side(border_spec.get("left"), side)
    right = _build_side(border_spec.get("right"), side)

    return Border(top=top, bottom=bottom, left=left, right=right)


def _build_side(side_spec, default_side):
    """构建单个边框 Side"""
    if not side_spec:
        return default_side
    style = BORDER_STYLE_MAP.get(side_spec.get("style", "thin"), "thin")
    color = _parse_color(side_spec.get("color", "#000000")) or "FF000000"
    return Side(style=style, color=color)


def _resolve_number_format(fmt_spec):
    """解析数字格式"""
    if not fmt_spec:
        return None
    if isinstance(fmt_spec, str):
        if fmt_spec in DEFAULT_NUMBER_FORMAT:
            return DEFAULT_NUMBER_FORMAT[fmt_spec]
        return fmt_spec
    return None


def _apply_style_to_cell(cell, style_spec, default_font, default_align, warnings):
    """对单个单元格应用样式"""
    font_spec = style_spec.get("font") if style_spec else None
    fill_spec = style_spec.get("fill") if style_spec else None
    align_spec = style_spec.get("alignment") if style_spec else None
    border_spec = style_spec.get("border") if style_spec else None
    number_fmt = style_spec.get("number_format") if style_spec else None

    font = _build_font(font_spec, default_font)
    if font:
        cell.font = font

    fill = _build_fill(fill_spec)
    if fill:
        cell.fill = fill

    alignment = _build_alignment(align_spec, default_align)
    if alignment:
        cell.alignment = alignment

    border = _build_border(border_spec)
    if border:
        cell.border = border

    nf = _resolve_number_format(number_fmt)
    if nf:
        cell.number_format = nf


def _apply_style_to_range(ws, range_spec, style_spec, default_font, default_align, warnings):
    """对范围内的所有单元格应用样式"""
    rng = _resolve_range(range_spec, ws, warnings)
    if rng is None:
        return

    min_col, min_row, max_col, max_row = rng
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            _apply_style_to_cell(cell, style_spec, default_font, default_align, warnings)


def _apply_conditional_format(ws, cf_spec, warnings):
    """应用条件格式"""
    range_str = cf_spec.get("range")
    if not range_str:
        warnings.append("条件格式缺少 range 参数")
        return

    rng = _resolve_range(range_str, ws, warnings)
    if rng is None:
        return

    min_col, min_row, max_col, max_row = rng
    range_addr = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

    cf_type = cf_spec.get("type", "cell_value")

    if cf_type == "color_scale":
        _apply_color_scale(ws, range_addr, cf_spec, warnings)
    elif cf_type == "data_bar":
        _apply_data_bar(ws, range_addr, cf_spec, warnings)
    elif cf_type == "icon_set":
        _apply_icon_set(ws, range_addr, cf_spec, warnings)
    elif cf_type == "formula":
        _apply_formula_rule(ws, range_addr, cf_spec, warnings)
    else:
        _apply_cell_value_rule(ws, range_addr, cf_spec, warnings)


def _apply_cell_value_rule(ws, range_addr, cf_spec, warnings):
    """应用基于单元格值的条件格式"""
    operator_raw = cf_spec.get("operator", "greaterThan")
    operator = CONDITIONAL_OPERATOR_MAP.get(operator_raw, operator_raw)

    formula = cf_spec.get("formula")
    value = cf_spec.get("value")

    style_spec = cf_spec.get("style", {})
    font = _build_font(style_spec.get("font"))
    fill = _build_fill(style_spec.get("fill"))
    border = _build_border(style_spec.get("border"))

    if formula:
        rule = CellIsRule(
            operator=operator,
            formula=[formula],
            font=font,
            fill=fill,
            border=border,
        )
    elif value is not None:
        if operator in ("between", "notBetween") and isinstance(value, list) and len(value) == 2:
            rule = CellIsRule(
                operator=operator,
                formula=[str(value[0]), str(value[1])],
                font=font,
                fill=fill,
                border=border,
            )
        else:
            rule = CellIsRule(
                operator=operator,
                formula=[str(value)],
                font=font,
                fill=fill,
                border=border,
            )
    else:
        warnings.append(f"条件格式缺少 value 或 formula 参数")
        return

    ws.conditional_formatting.add(range_addr, rule)


def _apply_color_scale(ws, range_addr, cf_spec, warnings):
    """应用色阶条件格式"""
    color_scale = cf_spec.get("color_scale", {})
    rule = ColorScaleRule(
        start_type=color_scale.get("start_type", "min"),
        start_color=color_scale.get("start_color", "FF63BE7B"),
        mid_type=color_scale.get("mid_type"),
        mid_color=color_scale.get("mid_color"),
        end_type=color_scale.get("end_type", "max"),
        end_color=color_scale.get("end_color", "FFFF5959"),
    )
    ws.conditional_formatting.add(range_addr, rule)


def _apply_data_bar(ws, range_addr, cf_spec, warnings):
    """应用数据条条件格式"""
    bar_color = cf_spec.get("bar_color", "FF638EC6")
    rule = DataBarRule(
        start_type="min",
        end_type="max",
        color=bar_color,
    )
    ws.conditional_formatting.add(range_addr, rule)


def _apply_icon_set(ws, range_addr, cf_spec, warnings):
    """应用图标集条件格式"""
    icon_style = cf_spec.get("icon_style", "3Arrows")
    icon_style = ICON_SET_MAP.get(icon_style, icon_style)
    rule = IconSetRule(
        icon_style=icon_style,
        type="percent",
        values=[0, 33, 67],
    )
    ws.conditional_formatting.add(range_addr, rule)


def _apply_formula_rule(ws, range_addr, cf_spec, warnings):
    """应用公式条件格式"""
    formula = cf_spec.get("formula")
    if not formula:
        warnings.append("公式条件格式缺少 formula 参数")
        return

    style_spec = cf_spec.get("style", {})
    font = _build_font(style_spec.get("font"))
    fill = _build_fill(style_spec.get("fill"))
    border = _build_border(style_spec.get("border"))

    rule = FormulaRule(
        formula=[formula],
        font=font,
        fill=fill,
        border=border,
    )
    ws.conditional_formatting.add(range_addr, rule)


def _add_chart(ws, chart_spec, warnings):
    """添加图表"""
    chart_type_raw = chart_spec.get("type", "bar")
    chart_type_raw = CHART_ALIAS_MAP.get(chart_type_raw, chart_type_raw)

    if chart_type_raw not in CHART_TYPE_MAP:
        warnings.append(f"不支持的图表类型 '{chart_type_raw}'，已降级为柱状图")
        chart_type_raw = "bar"

    chart_key, chart_class = CHART_TYPE_MAP[chart_type_raw]
    chart = chart_class()

    data_range = chart_spec.get("data_range")
    if not data_range:
        warnings.append("图表缺少 data_range 参数")
        return

    rng = _resolve_range(data_range, ws, warnings)
    if rng is None:
        return

    min_col, min_row, max_col, max_row = rng

    categories_title = chart_spec.get("categories_title")
    if categories_title:
        cat_col = _find_column_by_header(ws, categories_title)
        if cat_col:
            min_col = cat_col

    data = Reference(ws, min_col=min_col, min_row=min_row,
                     max_col=max_col, max_row=max_row)

    cats = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    title = chart_spec.get("title")
    if title:
        chart.title = title

    x_axis_title = chart_spec.get("x_axis_title")
    if x_axis_title:
        chart.x_axis.title = x_axis_title

    y_axis_title = chart_spec.get("y_axis_title")
    if y_axis_title:
        chart.y_axis.title = y_axis_title

    legend_pos = chart_spec.get("legend_position", "right")
    legend_map = {
        "bottom": "b",
        "top": "t",
        "left": "l",
        "right": "r",
        "底部": "b",
        "顶部": "t",
        "左侧": "l",
        "右侧": "r",
    }
    chart.legend.position = legend_map.get(legend_pos, "r")

    style = chart_spec.get("style")
    if style:
        try:
            chart.style = int(style)
        except (ValueError, TypeError):
            pass

    show_labels = chart_spec.get("show_data_labels", False)
    if show_labels:
        try:
            if chart_type_raw in ("pie", "pie3d"):
                chart.dataLabels = DataLabelList()
                chart.dataLabels.showPercent = True
                chart.dataLabels.showCatName = True
            else:
                for series in chart.series:
                    series.dLbls = DataLabelList()
                    series.dLbls.showVal = True
        except Exception:
            pass

    position = chart_spec.get("position", "E2")
    pos_match = re.match(r'^([A-Z]+)(\d+)$', position.strip().upper())
    if pos_match:
        anchor_col = pos_match.group(1)
        anchor_row = int(pos_match.group(2))
    else:
        anchor_col = "E"
        anchor_row = 2

    width = chart_spec.get("width_cm", 15)
    height = chart_spec.get("height_cm", 10)

    ws.add_chart(chart, f"{anchor_col}{anchor_row}")


def _setup_sheet(ws, sheet_spec, warnings):
    """设置单个工作表"""
    sheet_name = sheet_spec.get("name", ws.title)
    ws.title = sheet_name[:31]

    tab_color = sheet_spec.get("tab_color")
    if tab_color:
        color = _parse_color(tab_color)
        if color:
            ws.sheet_properties.tabColor = color

    default_font = dict(DEFAULT_FONT)
    if sheet_spec.get("default_font"):
        default_font.update(sheet_spec["default_font"])

    default_align = dict(DEFAULT_ALIGNMENT)
    if sheet_spec.get("default_alignment"):
        default_align.update(sheet_spec["default_alignment"])

    # 写入数据
    data = sheet_spec.get("data", [])
    for r_idx, row_data in enumerate(data, 1):
        for c_idx, value in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 应用样式
    styles = sheet_spec.get("styles", [])
    for style_def in styles:
        range_spec = style_def.get("range")
        style_spec = style_def.get("style", style_def)
        if range_spec:
            _apply_style_to_range(ws, range_spec, style_spec, default_font, default_align, warnings)

    # 合并单元格
    merged_cells = sheet_spec.get("merged_cells", [])
    for mc in merged_cells:
        mc_range = mc.get("range")
        if mc_range:
            try:
                ws.merge_cells(mc_range)
                content = mc.get("content")
                if content:
                    rng = _parse_range(mc_range)
                    if rng:
                        cell = ws.cell(row=rng[1], column=rng[0])
                        cell.value = content
                mc_style = mc.get("style")
                if mc_style:
                    _apply_style_to_range(ws, mc_range, mc_style, default_font, default_align, warnings)
            except Exception as e:
                warnings.append(f"合并单元格 {mc_range} 失败: {str(e)}")

    # 列宽
    columns = sheet_spec.get("columns", [])
    for i, col_spec in enumerate(columns, 1):
        col_letter = get_column_letter(i)
        if isinstance(col_spec, dict):
            width = col_spec.get("width")
            if width:
                ws.column_dimensions[col_letter].width = width
            hidden = col_spec.get("hidden", False)
            if hidden:
                ws.column_dimensions[col_letter].hidden = True
        elif isinstance(col_spec, (int, float)):
            ws.column_dimensions[col_letter].width = col_spec

    # 行高
    rows_spec = sheet_spec.get("rows", [])
    for i, row_spec in enumerate(rows_spec, 1):
        if isinstance(row_spec, dict):
            height = row_spec.get("height")
            if height:
                ws.row_dimensions[i].height = height
            hidden = row_spec.get("hidden", False)
            if hidden:
                ws.row_dimensions[i].hidden = True
        elif isinstance(row_spec, (int, float)):
            ws.row_dimensions[i].height = row_spec

    # 自动调整列宽
    auto_fit = sheet_spec.get("auto_fit_columns", False)
    if auto_fit:
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value:
                    # 中文字符按2个宽度计算
                    val = str(cell.value)
                    length = 0
                    for ch in val:
                        if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
                            length += 2
                        else:
                            length += 1
                    max_length = max(max_length, length)
            ws.column_dimensions[col_letter].width = min(max_length + 4, 60)

    # 冻结窗格
    freeze = sheet_spec.get("freeze_pane")
    if freeze:
        ws.freeze_panes = freeze

    # 条件格式
    conditional_formats = sheet_spec.get("conditional_formats", [])
    for cf in conditional_formats:
        _apply_conditional_format(ws, cf, warnings)

    # 图表
    charts = sheet_spec.get("charts", [])
    for chart_spec in charts:
        _add_chart(ws, chart_spec, warnings)


def create_excel_workbook(filename, headers=None, rows=None, formatting=None, output_dir=None):
    """创建格式化 Excel 工作簿

    Args:
        filename: 文件名（不含扩展名）
        headers: 简单模式的表头字符串（逗号分隔），向后兼容
        rows: 简单模式的数据行字符串（换行+逗号分隔），向后兼容
        formatting: 排版描述 JSON 对象

    Returns:
        str: 结果消息，包含文件路径和可能的警告
    """
    if output_dir is None:
        output_dir = os.path.join("document_output", "excel_output")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"{filename}.xlsx")

    warnings = []
    wb = Workbook()

    if formatting and isinstance(formatting, dict):
        sheets = formatting.get("sheets", [])
        if not sheets:
            warnings.append("未指定工作表(sheets)，生成了空白工作簿")

        for idx, sheet_spec in enumerate(sheets):
            if idx == 0:
                ws = wb.active
            else:
                ws = wb.create_sheet()

            _setup_sheet(ws, sheet_spec, warnings)

        # 调整 sheet 顺序
        sheet_order = formatting.get("sheet_order")
        if sheet_order:
            try:
                wb._sheets = [wb[name] for name in sheet_order if name in wb.sheetnames]
            except Exception:
                pass
    else:
        # 简单模式（向后兼容）
        ws = wb.active
        ws.title = filename[:31]

        header_font = Font(bold=True, size=12, color="FFFFFFFF")
        header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        header_list = []
        if headers:
            header_list = [h.strip() for h in headers.split(",")]
            for col_idx, header in enumerate(header_list, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

        row_list = []
        if rows:
            row_list = [r.strip() for r in rows.strip().split("\n") if r.strip()]
            for row_idx, row_str in enumerate(row_list, 2):
                cols = [c.strip() for c in row_str.split(",")]
                for col_idx, value in enumerate(cols, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

        # 自动调整列宽
        for col_idx in range(1, max(len(header_list), 1) + 1):
            max_length = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        val = str(cell.value)
                        length = 0
                        for ch in val:
                            if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
                                length += 2
                            else:
                                length += 1
                        max_length = max(max_length, length)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 6, 60)

    wb.save(filepath)

    result = f"Excel文件已成功保存至: {filepath}"
    if warnings:
        result += "\n[提示] " + "; ".join(warnings)
    return result